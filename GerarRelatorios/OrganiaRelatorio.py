from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment

def tirarMesclagem(aba_ativa):
    for items in sorted(aba_ativa.merged_cell_ranges):
        aba_ativa.unmerge_cells(str(items))

def tiraFormatoDasCelulas(aba_ativa):

    for row in aba_ativa.iter_rows():
        for celula in row:
            celula.style = "Normal"
            celula.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def limpa(aba_ativa):
    #Limpando Cabeçalhos
    aba_ativa.delete_rows(1,3)

    #Limpando colunas a mais
    aba_ativa.delete_cols(2,7)

def deletandoLinhasLixo(aba_ativa, lixo,cabecalho):
    for celula in aba_ativa["A"]:
        if celula.value == cabecalho[0]:
            linha = celula.row
            aba_ativa.delete_rows(linha,1)
        if celula.value == lixo:
            linha = celula.row
            aba_ativa.delete_rows(linha,2)


def deletandoLinhasEmBranco(aba_ativa): # -- NÂO ESTA FUNCIONANDO
    for celula in aba_ativa["A"]:
        if celula.row == None or celula.row == " ":
            #linha = celula.row
            aba_ativa.delete_rows(celula.row,1) 

def inserirLinhasParaCabeçalho(aba_ativa):
    aba_ativa.insert_rows(1,2)


def inserindoCabeçalho(aba_ativa,cabecalho):
    i=0
    for celula in aba_ativa["1"]:
        linha = cabecalho[i]
        celula.value = linha
        celula.font = Font(bold=True, size=12)
        i = i+1

def definindoLarguraDasColunas(aba_ativa):
    aba_ativa.column_dimensions["A"].width = 24
    aba_ativa.column_dimensions["B"].width = 16
    aba_ativa.column_dimensions["C"].width = 20
    aba_ativa.column_dimensions["D"].width = 13
    aba_ativa.column_dimensions["E"].width = 20
    aba_ativa.column_dimensions["F"].width = 21
    aba_ativa.column_dimensions["G"].width = 25
    aba_ativa.column_dimensions["H"].width = 15
    aba_ativa.column_dimensions["I"].width = 27

def centralizandoCabeçalho(aba_ativa):
    aba_ativa["A1"].alignment = Alignment(horizontal='center',vertical='center', wrap_text=True)
    aba_ativa["B1"].alignment = Alignment(horizontal='center',vertical='center', wrap_text=True)
    aba_ativa["C1"].alignment = Alignment(horizontal='center',vertical='center', wrap_text=True)
    aba_ativa["D1"].alignment = Alignment(horizontal='center',vertical='center', wrap_text=True)
    aba_ativa["E1"].alignment = Alignment(horizontal='center',vertical='center', wrap_text=True)
    aba_ativa["F1"].alignment = Alignment(horizontal='center',vertical='center', wrap_text=True)
    aba_ativa["G1"].alignment = Alignment(horizontal='center',vertical='center', wrap_text=True)
    aba_ativa["H1"].alignment = Alignment(horizontal='center',vertical='center', wrap_text=True)
    aba_ativa["I1"].alignment = Alignment(horizontal='center',vertical='center', wrap_text=True)

def definindoAlturaDasLinhas(aba_ativa):
    i=2
    for celula in aba_ativa["A"]:
        aba_ativa.row_dimensions[i].height = 80
        i = i+1

    aba_ativa.row_dimensions[1].height = 30
    aba_ativa.row_dimensions[2].height = 10


#-------------------------------------------------------            
planilha = load_workbook("Pasta1.xlsx")
aba_ativa = planilha.active

cabecalho = ["Processo","Caracteristica","Orgão Julgador", "Autuado em", "Classe Julgador",
"Polo Ativo", "Polo Passivo", "Nó(s) Atual(is)", "Ultimo Movimento"]
lixo = '««'

tirarMesclagem(aba_ativa)
tiraFormatoDasCelulas(aba_ativa)
limpa(aba_ativa)        
deletandoLinhasLixo(aba_ativa, lixo, cabecalho)
#deletandoLinhasEmBranco(aba_ativa)    
inserirLinhasParaCabeçalho(aba_ativa)
inserindoCabeçalho(aba_ativa, cabecalho)
definindoLarguraDasColunas(aba_ativa)
centralizandoCabeçalho(aba_ativa)
definindoAlturaDasLinhas(aba_ativa)


#Salvando planilha modificada
planilha.save("PlanilhaModificada.xlsx")
